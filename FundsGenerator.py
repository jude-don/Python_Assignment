from datetime import datetime
import time

username = input("Hi, I'm Elsie, your time and wage assistant. What would you like me to call you?\n").title()
print("Nice to meet you %s!\n" %username)
print("I will help you know how long you worked and how much you earned for a task you're about to start or an already completed one.\n") 


# algorithm to obtain start date and time
task_name = input("Please enter the name of the task: \n")#specify name to easily identify task.
PayRate = int(input("Please enter the amount of money paid per hour in figures: \n"))
answer = input("Are you starting now?Please type Y for yes and N for no. \n")#ask if user is starting now
if answer == "Y" or answer == "y": # allow room for capitalization error
    print("All the best %s!" %username)
    # Setting system datetime as start datetime
    dt = datetime.now()
    first_year = dt.year
    first_month = dt.month
    first_day = dt.day
    first_hour = dt.hour
    first_minute = dt.minute
    while True:
        stop_button = input("When you want to end. Please enter the keyword Elsie \n")
        if stop_button == "Elsie" or stop_button == "elsie":
            break
        else:
            print("%s are you trying to end" %username)
            continue

    #Setting system datetime as end datetime              
    st = datetime.now()
    second_year = st.year
    second_month = st.month
    second_day = st.day
    second_hour = st.hour
    second_minute = st.minute
    
else:
    print("Oh! you already completed the task? Good work!\nI'll tell you how long you worked and how much you earned.")

    ###########################################################################################
    #function to get date
    def get_date(description):
    #validate year
        while True:
            year = input("Please enter the year of the %s\n" % (description))
            if year.isnumeric():
                break
            else:
                print("your input is invalid. Try again! Enter a number this time\n")
                continue
    #validate month
        while True:
            month = input("Please enter the month of the %s, should be between 1-12: \n" % (description))
            if month.isnumeric():
                if int(month) >= 1 and int(month) <= 12:
                    break
                else:
                    print("your input is invalid. Try again!\n")
                continue
            else:
                print("your input is invalid. Try again! Enter a number this time.\n")
                continue
            break
    #get and validate day            
        while True:
            if int(month) == 2 and int(year)%4 == 0:#if it's February in a leap year, we'd expect an input a number between 1 and 29
                day = input("Please enter day of %s, remember this year february has 29 days: \n"% (description)) 
                if day.isnumeric():
                    if int(day) >= 1 and int(day) <= 29:
                        break
                    else:
                        print("Your input is invalid, Try again!\n")
                    continue
                else:
                    print("your input is invalid. Try again! Enter a number this time.\n")
                    continue

            elif int(month) == 2 and int(year)%4 != 0:#if it's February in a regular year, we'd expect an input a number between 1 and 28
                day = input("Please enter day of %s, remember february has 28 days: \n"% (description))
                if day.isnumeric():
                    if int(day) >= 1 and int(day) <= 28:
                       break
                    else:
                        print("Your input is invalid, Try again!\n")
                    continue
                else:
                    print("your input is invalid. Try again! Enter a number this time.\n")
                    continue

            elif month in [9,4,6,11]:# September, April, June and November have 30 days
                day = input("Please enter day of %s. should be between 1 and 30: \n"% (description))
                if day.isnumeric():
                    if int(day) >= 1 and int(day) <= 30:
                        break
                    else:
                        print("Your input is invalid, Try again!\n")
                    continue
                else:
                    print("your input is invalid. Try again! Enter a number this time.\n")
                    continue
                
            else:# All the rest are 31
                day = input("Please enter day of %s. should be between 1 and 31: \n"% (description))
                if day.isnumeric():
                    if int(day) >= 1 and int(day) <= 31:
                        break
                    else:
                        print("Your input is invalid, Try again!\n")
                    continue
                else:
                    print("your input is invalid. Try again! Enter a number this time.\n")
                    continue
        date = [year,month,day] 
        return date    
    #end of function
    #####################################################################################################

    try:
        # algorithm to obtain start date and time
        description = "start date"#Argument for get_date() function.

        start_date = get_date(description)# get start date using pre-defined function

        first_year = int(start_date[0])
        first_month = int(start_date[1])
        first_day = int(start_date[2])


        first_time = input("Please enter time of start date in 24 hour format, separated by commas. For example 12hours 30 minutes as 12,30: \n")
        mylist = first_time.split(',')
        first_hour = int(mylist[0])
        first_minute = int(mylist[1])

        # algorithm to obtain end date and time
        description = "end date"#Argument for get_date() function.
            
        end_date = get_date(description)

        second_year = int(end_date[0])
        second_month = int(end_date[1])
        second_day = int(end_date[2])

        second_time = input("Please enter time of end date in 24 hour format, separated by commas. For example 12hours 30 minutes as 12,30: \n")
        mylist1 = second_time.split(',')
        second_hour = int(mylist1[0])
        second_minute = int(mylist1[1])

    except IndexError:
        print("You entered an invalid input")

# Algorithm to find time difference and hours
first_date = datetime(first_year,first_month,first_day,first_hour,first_minute)

second_date = datetime(second_year,second_month,second_day,second_hour,second_minute)

difference_between_datetimes = second_date - first_date
hours_calculated = difference_between_datetimes.total_seconds()/3600
total_amount_received = hours_calculated * PayRate
currency = "$"
print('Great work %s! You spent %f minutes, which is %f %s on the task. And you earned %s%f\n' % (username,hours_calculated*60,hours_calculated,'hours',currency,total_amount_received))
print('Your bill is $200')
time.sleep(5)
print("Don't mind me, I'm just kidding.\nBye %s, hope to see you soon!" %username)


# Converting to an Excel file
import openpyxl
print("Please choose create or exisiting per the instructions in the next line. However, if you did not create the excel file with this program please do not choose exisitng for simplicity.")
decision = input("Please type Create if you want to create a new excel file or type Exisitng if you want to use an existing file. \n")
if decision == "Create" or decision == "create":
    filename = input("What do you want to save the file as \n")
    wb = openpyxl.Workbook()
    sheet = wb.active
    A1 = sheet.cell(row = 1, column = 1)
    B1 = sheet.cell(row = 1, column = 2)
    C1 = sheet.cell(row = 1, column = 3)
    D1 = sheet.cell(row = 1, column = 4)
    E1 = sheet.cell(row = 1, column = 5)
    F1 = sheet.cell(row = 1, column = 6)  
    G1 = sheet.cell(row = 1, column = 7)        
    A1.value = 'TaskName'
    B1.value = 'StartDate'
    C1.value = 'StartTime'
    D1.value = 'EndDate'
    E1.value = 'EndTime'
    F1.value = 'HoursSpent'
    G1.value = 'AmountReceived'
    A2 = sheet.cell(row = 2, column = 1)
    B2 = sheet.cell(row = 2, column = 2)
    C2 = sheet.cell(row = 2, column = 3)
    D2 = sheet.cell(row = 2, column = 4)
    E2 = sheet.cell(row = 2, column = 5)
    F2 = sheet.cell(row = 2, column = 6)  
    G2 = sheet.cell(row = 2, column = 7)
    A2.value = task_name
    B2.value = str(first_year)+"-"+str(first_month)+"-"+str(first_day)
    C2.value = str(first_hour)+":"+str(first_minute)
    D2.value = str(second_year)+"-"+str(second_month)+"-"+str(second_day)
    E2.value = str(second_hour)+":"+str(second_minute)
    F2.value = str(hours_calculated)
    G2.value = "$" + str(total_amount_received)
    filename = filename + ".xlsx"
    wb.save(filename)
else:
    filename = input("Please enter the excel filename with the .xlsx extension. eg, Demo.xlsx \n")
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    rowcount = sheet.max_row # this shows the number of populated rows
    rtsw = rowcount + 1  # this is the row to start with
    A3 = sheet.cell(row = rtsw, column = 1)
    B3 = sheet.cell(row = rtsw, column = 2)
    C3 = sheet.cell(row = rtsw, column = 3)
    D3 = sheet.cell(row = rtsw, column = 4)
    E3 = sheet.cell(row = rtsw, column = 5)
    F3 = sheet.cell(row = rtsw, column = 6)  
    G3 = sheet.cell(row = rtsw, column = 7)
    A3.value = task_name
    B3.value = str(first_year)+"-"+str(first_month)+"-"+str(first_day)
    C3.value = str(first_hour)+":"+str(first_minute)
    D3.value = str(second_year)+"-"+str(second_month)+"-"+str(second_day)
    E3.value = str(second_hour)+":"+str(second_minute)
    F3.value = str(hours_calculated)
    G3.value = "$" + str(total_amount_received)
    book.save(filename)
print('Your bill is $200')
time.sleep(2.2)
print("Don't mind me, I'm just kidding.\n Bye %s, hope to see you soon!" %username)
